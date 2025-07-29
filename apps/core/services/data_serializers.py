import csv
import json
from io import StringIO, BytesIO, TextIOWrapper
from typing import Dict, List, Any, Optional
from abc import ABC, abstractmethod
from django.utils import timezone
import zipfile


class BaseDataSerializer(ABC):
    @abstractmethod
    def serialize(self, data: Dict[str, List[Dict[str, Any]]], metadata: Dict[str, Any]) -> bytes:
        pass


class CSVDataSerializer(BaseDataSerializer):
    def serialize(self, data: Dict[str, List[Dict[str, Any]]], metadata: Dict[str, Any]) -> bytes:
        output = BytesIO()
        
        for table_name, records in data.items():
            if not records:
                continue
                
            csv_output = StringIO()
            if records:
                fieldnames = records[0].keys()
                writer = csv.DictWriter(csv_output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(records)
            
            output.write(f"=== {table_name.upper()} ===\n".encode('utf-8'))
            output.write(csv_output.getvalue().encode('utf-8'))
            output.write(b"\n\n")
        
        output.seek(0)
        return output.getvalue()


class ZIPDataSerializer(BaseDataSerializer):
    def serialize(self, data: Dict[str, List[Dict[str, Any]]], metadata: Dict[str, Any]) -> bytes:
        output = BytesIO()
        zip_buffer = zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED)
        
        for table_name, records in data.items():
            if not records:
                continue
                
            csv_output = StringIO()
            fieldnames = records[0].keys()
            writer = csv.DictWriter(csv_output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
            
            zip_buffer.writestr(f"{table_name}.csv", csv_output.getvalue())
        
        metadata_output = StringIO()
        metadata_writer = csv.DictWriter(metadata_output, fieldnames=['key', 'value'])
        metadata_writer.writeheader()
        for key, value in metadata.items():
            metadata_writer.writerow({'key': key, 'value': str(value)})
        
        zip_buffer.writestr("metadata.csv", metadata_output.getvalue())
        zip_buffer.close()
        
        output.seek(0)
        return output.getvalue()


class ExcelDataSerializer(BaseDataSerializer):
    def serialize(self, data: Dict[str, List[Dict[str, Any]]], metadata: Dict[str, Any]) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        
        # Remover la hoja por defecto
        wb.remove(wb.active)
        
        # Crear hoja de metadatos
        metadata_ws = wb.create_sheet("Información General")
        metadata_ws.append(["Campo", "Valor"])
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for key, value in metadata.items():
            metadata_ws.append([key, str(value)])
        
        # Aplicar estilos a encabezados de metadatos
        for cell in metadata_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Crear hojas para cada tabla de datos
        for table_name, records in data.items():
            if not records:
                continue
                
            ws = wb.create_sheet(table_name.replace('_', ' ').title())
            
            if records:
                # Añadir encabezados
                headers = list(records[0].keys())
                ws.append(headers)
                
                # Aplicar estilos a encabezados
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Añadir datos
                for record in records:
                    row_data = []
                    for header in headers:
                        value = record.get(header, '')
                        # Convertir valores complejos a string
                        if value is None:
                            value = ''
                        elif not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                        row_data.append(value)
                    ws.append(row_data)
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class JSONDataSerializer(BaseDataSerializer):
    def serialize(self, data: Dict[str, List[Dict[str, Any]]], metadata: Dict[str, Any]) -> bytes:
        export_data = {
            'metadata': metadata,
            'data': data
        }
        return json.dumps(export_data, indent=2, default=str).encode('utf-8')


class DataSerializerFactory:
    _serializers = {
        'csv': CSVDataSerializer,
        'zip': ZIPDataSerializer,
        'excel': ExcelDataSerializer,
        'json': JSONDataSerializer,
    }
    
    @classmethod
    def get(cls, format_name: str) -> BaseDataSerializer:
        serializer_class = cls._serializers.get(format_name.lower())
        if not serializer_class:
            raise ValueError(f"Unsupported format: {format_name}")
        return serializer_class()
    
    @classmethod
    def get_available_formats(cls) -> List[str]:
        return list(cls._serializers.keys())
